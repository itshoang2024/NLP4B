from __future__ import annotations

from pathlib import Path
from typing import Optional, Union
from urllib.parse import parse_qs, urlparse

import openpyxl
import pandas as pd


PathLike = Union[str, Path]


def extract_youtube_video_id(url: str) -> Optional[str]:
    """
    Extract the YouTube video ID from a standard YouTube or youtu.be URL.

    Supported examples:
    - https://www.youtube.com/watch?v=oA-BhGNK7qw
    - https://youtu.be/oA-BhGNK7qw
    """
    if not isinstance(url, str) or not url.strip():
        return None

    parsed = urlparse(url.strip())
    netloc = parsed.netloc.lower()

    if "youtu.be" in netloc:
        return parsed.path.lstrip("/") or None

    if "youtube.com" in netloc:
        qs = parse_qs(parsed.query)
        values = qs.get("v", [])
        return values[0] if values else None

    return None


def choose_best_url_from_cell(cell) -> Optional[str]:
    """
    Return the real URL from an Excel cell.

    Priority:
    1. Hyperlink target (works for Google Sheets chips exported to Excel)
    2. Raw URL text in the cell value
    """
    if getattr(cell, "hyperlink", None) and getattr(cell.hyperlink, "target", None):
        target = str(cell.hyperlink.target).strip()
        return target or None

    value = cell.value
    if isinstance(value, str):
        value = value.strip()
        if value.startswith(("http://", "https://")):
            return value

    return None


def choose_title_from_cell(cell, url: Optional[str]) -> Optional[str]:
    """
    Return the visible text shown in the Excel cell when it differs from the raw URL.
    """
    value = cell.value
    visible_text = value.strip() if isinstance(value, str) else ""

    if visible_text and visible_text != (url or ""):
        return visible_text

    return None


def _build_manifest_rows(ws) -> list[dict]:
    rows: list[dict] = []

    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if not row:
            continue

        url_cell = row[0]
        url = choose_best_url_from_cell(url_cell)
        if not url:
            continue

        video_id = extract_youtube_video_id(url)
        title = choose_title_from_cell(url_cell, url)

        rows.append(
            {
                "row_number": idx - 1,
                "url": url,
                "title": title,
                "video_id": video_id,
            }
        )

    return rows


def load_manifest_from_excel(
    excel_path: PathLike,
    sheet_name: Union[int, str] = 0,
    deduplicate_video_ids: bool = False,
) -> pd.DataFrame:
    """
    Load a one-column Excel file whose first column must be `url`.

    Notes:
    - Extra columns are ignored.
    - The visible cell text is used as `title` when available.
    - Hyperlink targets are preserved via openpyxl.
    """
    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[sheet_name]] if isinstance(sheet_name, int) else wb[sheet_name]

    first_header = ws[1][0].value
    if first_header != "url":
        raise ValueError(f'Expected first column to be "url", but got: {first_header!r}')

    df = pd.DataFrame(_build_manifest_rows(ws))
    if df.empty:
        raise ValueError("No valid rows found in Excel.")

    if df["video_id"].isna().any():
        bad_rows = df[df["video_id"].isna()][["row_number", "url"]]
        raise ValueError(
            "Some rows do not have a valid YouTube video_id. Please check these rows:\n"
            + bad_rows.to_string(index=False)
        )

    if deduplicate_video_ids:
        df = df.drop_duplicates(subset=["video_id"], keep="first").reset_index(drop=True)

    df["local_video_path"] = df["video_id"].apply(lambda x: f"videos/{x}.mp4")
    df["info_json_path"] = df["video_id"].apply(lambda x: f"info_json/{x}.info.json")

    return df


def save_normalized_manifest(df: pd.DataFrame, output_csv_path: PathLike) -> Path:
    """
    Save the normalized manifest as CSV.
    """
    output_csv_path = Path(output_csv_path)
    output_csv_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(output_csv_path, index=False)
    return output_csv_path
